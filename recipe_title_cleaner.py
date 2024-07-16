import pathlib
import textwrap

import google.generativeai as genai
from openpyxl import load_workbook 

genai.configure(api_key="AIzaSyA0tS2xolgfP_R9f2_BvCdaQBZYYORmJUk")
model = genai.GenerativeModel('gemini-pro')

def read_and_combine_excel_entries(start_index):
    file_path="Recipe_Title_Indian.xlsx"
    combined_string = ''
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active
    
    # Read the next 50 entries in the first column starting from start_index
    for i in range(start_index + 1, start_index + 51):
        cell_value = sheet.cell(row=i, column=1).value
        if cell_value is not None:  # Ensure the cell is not empty
            combined_string += str(cell_value)  # Append the first column entry to the string
    
    workbook.close()
    return combined_string


#training promt
promt0="Masala chole bhatura ravi masala chole bhatura dosa idli samosa raghu samosa i have given u a sample of dish names, u have to remove the common one and any chef's name or any other unnecessary thing like dish numnber or repitions and give unique dish name, for eg; ravi dosa and aman dosa are same and is just dosa"
prompt1="""T and T Famous Chicken Curry should be just chicken curry as you should remove any restaurant name from recipe and if we  see  any recipe which we have already have in our list remove it then  and Murag on Mondays should be just murag as remove any name of day or year or date from recipe name as well and My 1000th Recipe - Mango Paneer should be just mango curry as you remove any number in name of recipe as well also Bob's Carrot Halva should be just carrot halwa as you remove name of chef as well and Yogi Bhajan's Onion Soup should be just Onion soup as you remove  any chef name and Yummy Patties should be just patties as you remove any taste adjective  and similaryly fix these Madhur Jaffrey's Prawn (Shrimp) Curry
Julie's Baked Chicken Korma & Rice
Kori's Roti (West Indian Chicken Curry) """

prompt2 ="""here are more examples Shelly's Meat Curry -> Meat Curry 
Authentic Bangladeshi Beef Curry -> Bangladeshi Beef Curry
Indian-Style Vegetable Rice -> Indian Vegetable Rice 
Vegetable Biryani (Tehri) -> Vegetable Biryani 
Shyam's Goat Biryani -> Goat Biryani
Crazy Spicy Chicken -> Spicy Chicken 
Slow-Cooked Chicken Biryani -> Chicken Biryani 
Easy Chicken Phaal -> Chicken Phaal 
Chicken Chicken Curry -> Chicken Curry 
Wendy's Indian Butter Chicken -> Butter Chicken 
Authentic South Indian Biryani -> South Indian Biryani 
Easy Chicken Curry -> Chicken Curry 
Delicious and Fast Chicken Curry -> Chicken Curry 
Chef John's Tandoori Chicken -> Tandoori Chicken 
Chef John's Chicken Tikka Masala -> Chicken Tikka Masala 
Slow Cooker Butter Chicken -> Butter Chicken ()
Indian-Style "Chettinad" Chicken -> Chettinad-Style Chicken
Traditional Chicken Curry -> Chicken Curry 
Chicken Korma II -> Chicken Korma 
Easy Slow Cooker Chicken Tikka Masala -> Chicken Tikka Masala 
Fried Chicken Biryani (Filipino-Style) -> Chicken Biryani 
also delete if there are any repitions"""
promt3="""i am teaching u the protocols of giving unique names one by one, then u have to do this for a large sample of data u have to also use your knowledge to figure out similardish names from the list and remove them first report the modified names or delelted names then print all the with this heading unique names in excel format and don't print anything after excel names lets start with following recipe titles Spicy Crab Curry - Bangla Style
Machhere Jhol (Bengali Fish Curry)
Egg and Potato Curry
Butter Lamb Gravy
Beef Samosas
Spicy Shrimp (Chingri Maach)
Gulab Jamun or Kala Jam (Waffle Balls)
Prawn Malai Curry
Shelly's Meat Curry
Indian Fish Curry
Adipoli Parathas
Spicy Potato Noodles (Bataka Sev)
Authentic Bangladeshi Beef Curry
Indian Mustard Fish
Bengali Dhal
Bengali Chicken Curry with Potatoes
Vegetable Pulao
Jeera Fried Rice
Saffron Rice with Raisins and Cashews
should give following results 
Spicy Crab Curry (Bangla Style)
Machhere Jhol (Bengali Fish Curry)
Egg and Potato Curry
Butter Lamb Gravy
Beef Samosas
Spicy Shrimp (Chingri Maach)
Gulab Jamun or Kala Jam (Waffle Balls)
Prawn Malai Curry
Meat Curry
Indian Fish Curry
Adipoli Parathas
Spicy Potato Noodles (Bataka Sev)
Bangladeshi Beef Curry
Indian Mustard Fish
Bengali Dhal
Bengali Chicken Curry with Potatoes
Vegetable Pulao
Jeera Fried Rice
Saffron Rice with Raisins and Cashews
"""
training_prompt =[promt0,prompt1,prompt2,promt3]
def train():
    for prompt in training_prompt:
        response=model.generate_content(prompt)
        print(response.text)
        print("="*50)
#train()
# response = model.generate_content("What is the meaning of life?")
# print(response.text)

def get_more_names(index ):
    prompt="now apply the same protocol for following more recipes and maintain the list of unique recipes"+read_and_combine_excel_entries(index)
    response=model.generate_content(prompt)
    #
    print(response.text)
    return
get_more_names(1)
